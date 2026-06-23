#!/usr/bin/env python3
"""
Build data.json for the Wellabe Rate Competitiveness & 2027 Sales Impact app.

Reads the two source workbooks in data_sources/ and emits data.json, a compact
columnar dataset consumed by the static front-end:

  * Granular Plan G rate cells (state x zip3 x age x gender) with Wellabe's rate
    (lowest Medico entity) and the six Big-6 carrier-bucket benchmark rates.
  * Baseline 2026 projected sales by state x month (the 2027 carry-forward base).
  * Metadata: state list, bucket labels, generation date.

Usage:  python3 build_data.py
"""
import json
import os
import statistics
from collections import defaultdict
from datetime import date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RATES_FILE = os.path.join(HERE, "data_sources", "Ranks_v_Sales_MS_June.xlsx")
SALES_FILE = os.path.join(HERE, "data_sources", "MS_Sales_Tracking_2026.xlsx")
OUT_FILE = os.path.join(HERE, "data.json")

# The three Medico legal entities make up "Wellabe" in the rate table.
MEDICO = {
    "Medico Insurance Company",
    "Medico Life and Health Insurance Company",
    "Medico Corp Insurance Company",
}

# The six Big-6 buckets. Order is fixed and mirrored in the front-end.
BUCKETS = ["UHC", "Humana", "Aetna", "Cigna", "MOO", "Blues"]
BUCKET_LABELS = {
    "UHC": "UnitedHealthcare / AARP",
    "Humana": "Humana",
    "Aetna": "Aetna",
    "Cigna": "Cigna",
    "MOO": "Mutual of Omaha",
    "Blues": "Blue Cross Blue Shield",
}


def classify_bucket(name: str) -> str | None:
    """Map a carrier name to one of the six Big-6 buckets, else None."""
    n = name.lower()
    if "mutual of omaha" in n or "united of omaha" in n or "omaha" in n \
            or "united world" in n:
        return "MOO"
    if "aarp" in n or "unitedhealthcare" in n or "healthspring" in n:
        return "UHC"  # HealthSpring is a Cigna brand but tagged with UHC group historically
    if "humana" in n:
        return "Humana"
    if "aetna" in n or "continental life" in n or "american continental" in n:
        return "Aetna"
    if "cigna" in n or "american retirement life" in n or "loyal american" in n:
        return "Cigna"
    if ("blue" in n or "anthem" in n or "highmark" in n or "wellmark" in n
            or "carefirst" in n or "regence" in n or "premera" in n
            or "florida blue" in n or "excellus" in n or "capital blue" in n):
        return "Blues"
    return None


def load_big6_names(wb) -> set[str]:
    """Carrier name-variants tagged 'Big Six' in the Lookup sheet."""
    names = set()
    lk = wb["Lookup"]
    for row in lk.iter_rows(min_row=2, values_only=True):
        name, tag = row[5], row[6]
        if name and tag == "Big Six":
            names.add(name.strip())
    return names


def build_rate_cells():
    wb = openpyxl.load_workbook(RATES_FILE, read_only=True, data_only=True)
    big6_names = load_big6_names(wb)
    ws = wb["Rates G"]

    # cell key -> {company: rate}; also remember a primary city per cell
    cell_rates = defaultdict(dict)
    cell_city = {}
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[1] is None:
            continue
        st, zip3, age, gender, city, comp, rate = (
            row[3], row[4], row[5], row[6], row[8], row[7], row[9],
        )
        if comp is None or rate is None or st is None:
            continue
        key = (st, str(zip3), str(age), gender)
        cell_rates[key][comp.strip()] = rate
        if city and key not in cell_city:
            cell_city[key] = city
    wb.close()

    cells = []
    for (st, zip3, age, gender), d in cell_rates.items():
        med = [d[m] for m in MEDICO if m in d]
        if not med:
            continue  # no Wellabe rate -> nothing to compare
        wellabe = round(min(med), 1)

        # Aggregate Big-6 carriers into the six buckets (mean per bucket).
        bucket_vals = defaultdict(list)
        for comp, rate in d.items():
            if comp in big6_names:
                b = classify_bucket(comp)
                if b:
                    bucket_vals[b].append(rate)
        bench = []
        for b in BUCKETS:
            if bucket_vals[b]:
                bench.append(round(statistics.mean(bucket_vals[b]), 1))
            else:
                bench.append(None)
        if not any(v is not None for v in bench):
            continue  # no Big-6 reference in this cell

        cells.append([
            st, zip3, int(age), "M" if gender == "Male" else "F",
            cell_city.get((st, zip3, age, gender), ""), wellabe, bench,
        ])
    return cells


def build_sales():
    """Per-state 2026 baseline from the Projection sheet's "Projected w action"
    block (Approved basis) — the company's planned-action projection (~$288M).
    Layout: a header row whose col 4 starts with 'Projected w act', followed by
    one row per state with col 4 == 'Total', col 5 = state, col 6 = annual,
    cols 7..18 = months 1..12."""
    wb = openpyxl.load_workbook(SALES_FILE, read_only=True, data_only=True)
    ws = wb["Projection"]
    rows = list(ws.iter_rows(values_only=True))

    hdr = None
    for i, r in enumerate(rows):
        if isinstance(r[4], str) and r[4].startswith("Projected w act"):
            hdr = i
            break
    if hdr is None:
        raise RuntimeError("Could not find 'Projected w action' block in Projection sheet")

    by_state = {}
    for i in range(hdr + 1, len(rows)):
        r = rows[i]
        if r[4] == "Total" and isinstance(r[5], str) and isinstance(r[6], (int, float)):
            months = [round(r[7 + m] or 0, 2) for m in range(12)]
            by_state[r[5]] = {"annual": round(r[6], 2), "months": months}
        elif r[5] is None and r[6] is None and i > hdr + 2:
            break
    wb.close()
    return by_state


def main():
    print("Reading rate cells ...")
    cells = build_rate_cells()
    print(f"  {len(cells)} rate cells")

    print("Reading sales baseline ...")
    sales = build_sales()
    total = sum(s["annual"] for s in sales.values())
    print(f"  {len(sales)} states, baseline total ${total:,.0f}")

    states = sorted(sales.keys())
    out = {
        "generated": date.today().isoformat(),
        "buckets": BUCKETS,
        "bucketLabels": BUCKET_LABELS,
        "months": ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
        "salesStates": states,
        # cell schema: [state, zip3, age, gender, city, wellabe, [6 bucket rates]]
        "cellSchema": ["state", "zip3", "age", "gender", "city", "wellabe", "big6"],
        "cells": cells,
        "sales": sales,
        "baselineTotal": round(total, 2),
        "salesBasis": "Approved (issued) premium",
        "salesLine": "Projected w action (company planned-action 2026 projection)",
    }
    with open(OUT_FILE, "w") as f:
        json.dump(out, f, separators=(",", ":"))
    size = os.path.getsize(OUT_FILE) / 1024
    print(f"Wrote {OUT_FILE} ({size:.0f} KB)")


if __name__ == "__main__":
    main()
